import frappe
from frappe import _

from crm.api.doc import get_assigned_users, get_fields_meta
from crm.fcrm.doctype.crm_form_script.crm_form_script import get_form_script


@frappe.whitelist()
def get_lead(name):
	lead = frappe.get_doc("CRM Lead", name)
	lead.check_permission("read")

	lead = lead.as_dict()

	lead["fields_meta"] = get_fields_meta("CRM Lead")
	lead["_form_script"] = get_form_script("CRM Lead")
	return lead

@frappe.whitelist()
def get_lead_visits(name):
	"""
	Get linked site visits for a lead.
	"""
	if not name:
		return []

	visits = frappe.get_all(
		"CRM Site Visit",
		filters={
			"reference_type": "CRM Lead",
			"reference_name": name
		},
		fields=[
			"name",
			"visit_date",
			"visit_type", 
			"status",
			"priority",
			"sales_person",
			"visit_purpose",
			"visit_summary",
			"planned_start_time",
			"planned_end_time",
			"check_in_time",
			"check_out_time",
			"total_duration",
			"lead_quality",
			"feedback",
			"next_steps",
			"follow_up_required",
			"follow_up_date",
			"potential_value",
			"probability_percentage"
		],
		order_by="visit_date desc"
	)

	return visits

@frappe.whitelist()
def get_meeting_data_sections(doctype):
	if not frappe.db.exists("CRM Fields Layout", {"dt": doctype, "type": "Side Data Bar"}):
		return []
	layout = frappe.get_doc("CRM Fields Layout", {"dt": doctype, "type": "Side Data Bar"}).layout

	if not layout:
		return []

	import json
	from crm.fcrm.doctype.crm_fields_layout.crm_fields_layout import get_field_obj, handle_perm_level_restrictions

	layout = json.loads(layout)

	not_allowed_fieldtypes = [
		"Tab Break",
		"Section Break",
		"Column Break",
	]

	fields = frappe.get_meta(doctype).fields
	fields = [field for field in fields if field.fieldtype not in not_allowed_fieldtypes]

	for section in layout:
		section["name"] = section.get("name") or section.get("label")
		for column in section.get("columns") if section.get("columns") else []:
			for field in column.get("fields") if column.get("fields") else []:
				field_obj = next((f for f in fields if f.fieldname == field), None)
				if field_obj:
					field_obj = field_obj.as_dict()
					handle_perm_level_restrictions(field_obj, doctype)
					column["fields"][column.get("fields").index(field)] = get_field_obj(field_obj)

	return layout

@frappe.whitelist()
def get_importable_fields(doctype):
	meta = frappe.get_meta(doctype)
	fields = []
	
	# Skip standard fields that shouldn't be imported manually
	skip_fields = [
		"name", "owner", "creation", "modified", "modified_by", 
		"docstatus", "idx", "lft", "rgt", "old_parent", 
		"status_change_log", "products", "sla"
	]
	
	for field in meta.fields:
		if field.fieldname in skip_fields:
			continue
		if field.fieldtype in ["Section Break", "Column Break", "Tab Break", "HTML", "Button", "Table"]:
			continue
		if field.hidden or field.read_only or field.fieldname == "status":
			continue
			
		fields.append({
			"fieldname": field.fieldname,
			"label": field.label,
			"reqd": field.reqd,
			"fieldtype": field.fieldtype,
			"options": field.options
		})
		
	return fields

@frappe.whitelist()
def download_lead_import_template(fields, file_type="csv"):
	import json
	if isinstance(fields, str):
		fields = json.loads(fields)
		
	if file_type == "csv":
		import csv
		from io import StringIO
		
		f = StringIO()
		writer = csv.writer(f)
		writer.writerow(fields)
		
		frappe.response['filecontent'] = f.getvalue()
		frappe.response['type'] = 'csv'
		frappe.response['filename'] = 'crm_lead_import_template.csv'
	elif file_type == "excel":
		try:
			import openpyxl
			from openpyxl.worksheet.datavalidation import DataValidation
			from io import BytesIO
			
			wb = openpyxl.Workbook()
			ws = wb.active
			ws.title = "Import Template"
			
			# Header row
			ws.append(fields)
			
			# Create a hidden worksheet for large lists (links)
			data_ws = wb.create_sheet("Data")
			data_ws.sheet_state = 'hidden'
			
			meta = frappe.get_meta("CRM Lead")
			
			for col_idx, fieldname in enumerate(fields, 1):
				field = meta.get_field(fieldname)
				if not field: continue
				
				dv_options = []
				if field.fieldtype == "Select" and field.options:
					dv_options = field.options.split("\n")
				elif field.fieldtype == "Link":
					# Fetch existing values for Link fields (Top 100)
					dv_options = frappe.get_all(field.options, limit=100, pluck="name")
				
				if dv_options:
					# Clean options (remove empty strings)
					dv_options = [str(o).strip() for o in dv_options if o]
					
					if dv_options:
						# For short lists, we can use literal string
						# Excel has a length limit for literal strings in DV
						options_str = '"{}"'.format(",".join(dv_options))
						
						if len(options_str) < 255:
							dv = DataValidation(type="list", formula1=options_str, allow_blank=True)
						else:
							# For longer lists, use the hidden worksheet
							start_row = 1
							for idx, opt in enumerate(dv_options):
								data_ws.cell(row=start_row + idx, column=col_idx, value=opt)
							
							col_letter = openpyxl.utils.get_column_letter(col_idx)
							formula = "Data!${}$1:${}${}".format(col_letter, col_letter, len(dv_options))
							dv = DataValidation(type="list", formula1=formula, allow_blank=True)
						
						col_letter = openpyxl.utils.get_column_letter(col_idx)
						dv.add("{col}2:{col}1000".format(col=col_letter))
						ws.add_data_validation(dv)
			
			output = BytesIO()
			wb.save(output)
			output.seek(0)
			
			frappe.response['filecontent'] = output.getvalue()
			frappe.response['type'] = 'binary'
			frappe.response['filename'] = 'crm_lead_import_template.xlsx'
		except ImportError:
			frappe.throw(_("openpyxl is not installed. Please install it to use Excel templates."))

@frappe.whitelist()
def parse_import_file(file, file_type):
	if not file:
		frappe.throw(_("No file provided"))
		
	import json
	rows = []
	
	# Get file content if it's a file object or a path
	if hasattr(file, 'file_url'):
		file_path = frappe.get_site_path(file.file_url.strip("/"))
		with open(file_path, "rb") as f:
			content = f.read()
	elif isinstance(file, str) and file.startswith("data:"):
		# Handle base64 encoded data (often from frontend)
		import base64
		header, encoded = file.split(",", 1)
		content = base64.b64decode(encoded)
	else:
		frappe.throw(_("Invalid file format"))
		
	if file_type == "csv":
		import csv
		from io import StringIO
		
		# Detect encoding
		decoded_content = content.decode('utf-8', errors='ignore')
		f = StringIO(decoded_content)
		reader = csv.DictReader(f)
		for row in reader:
			rows.append(row)
			
	elif file_type == "excel":
		import openpyxl
		from io import BytesIO
		
		wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
		ws = wb.active
		
		headers = []
		for cell in ws[1]:
			headers.append(cell.value)
			
		for row in ws.iter_rows(min_row=2, values_only=True):
			data = {}
			for idx, value in enumerate(row):
				if idx < len(headers):
					data[headers[idx]] = value
			rows.append(data)
			
	return rows

@frappe.whitelist()
def bulk_import_leads(data):
	import json
	if isinstance(data, str):
		data = json.loads(data)
		
	if not data:
		return {"success": False, "message": _("No data provided")}
		
	meta = frappe.get_meta("CRM Lead")
	valid_fields = [df.fieldname for df in meta.fields] + ["naming_series"]
	
	created_count = 0
	errors = []
	
	for row in data:
		try:
			# Filter row to include only valid fields
			filtered_row = {k: v for k, v in row.items() if k in valid_fields and v is not None and v != ""}
			
			# Skip empty rows
			if not filtered_row:
				continue

			# Set lead_owner if not provided
			if not filtered_row.get("lead_owner"):
				filtered_row["lead_owner"] = frappe.session.user
				
			doc = frappe.get_doc({
				"doctype": "CRM Lead",
				**filtered_row
			})
			doc.insert()
			
			# Add assignment
			from frappe.desk.form.assign_to import add as add_assignment
			add_assignment({
				"doctype": "CRM Lead",
				"name": doc.name,
				"assign_to": [frappe.session.user],
				"bulk_assign": True,
				"re_assign": True
			})
			
			created_count += 1
		except frappe.LinkValidationError as e:
			errors.append({"row": row, "error": str(e)})
		except Exception as e:
			frappe.log_error(frappe.get_traceback(), _("Lead Bulk Import Error"))
			errors.append({"row": row, "error": str(e)})
			
	return {
		"success": True,
		"created_count": created_count,
		"error_count": len(errors),
		"errors": errors[:10] # Return first 10 errors
	}
